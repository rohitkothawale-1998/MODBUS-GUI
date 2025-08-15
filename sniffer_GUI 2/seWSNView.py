#!/usr/bin/env python3
# REON Modbus GUI (Modbus RTU)
# pip install wxPython pymodbus pyserial openpyxl

import os
import wx
import wxSerialConfigDialog
import serial
import threading
import time
from dataclasses import dataclass
from typing import List, Dict, Optional
from serial.tools import list_ports

import pymodbus
from pymodbus.client import ModbusSerialClient

try:
    from pymodbus import FramerType      # pymodbus 3.x
    _HAS_FRAMER = True
    _FRAMER_KW = "framer=FramerType.RTU"
except Exception:
    _HAS_FRAMER = False                   # pymodbus 2.x
    _FRAMER_KW = 'method="rtu"'

# ──────────────────────────────────────────────────────────────────────────────
# Register table (decoders, scales, units)

@dataclass(frozen=True)
class Reg:
    name: str
    addr: int
    words: int
    codec: str      # ascii | u16 | s16 | u32 | s32
    scale: float
    unit: str

DEVICE_DATA: List[Reg] = [
    Reg("Serial #",         0xC780, 15, "ascii", 1,     ""),
    Reg("Inverter SN",      0xC78F, 10, "ascii", 1,     ""),
    Reg("Production Date",  0xC7A0,  4, "ascii", 1,     ""),
    Reg("Firmware Version", 0xC783,  1, "u16",   1,     ""),
    Reg("HW Version",       0xC784,  1, "u16",   1,     ""),
    Reg("Model Number",     0xC785,  1, "u16",   1,     ""),
    Reg("Manufacturer",     0xC786,  1, "u16",   1,     ""),
]

RUNTIME_DATA: List[Reg] = [
    Reg("AC Input Voltage",    0x756A, 1, "u16",  0.1,   "V"),
    Reg("AC Input Current",    0x756B, 1, "s16",  0.1,   "A"),
    Reg("AC Input Power",      0x7571, 1, "s16",  1,     "VA"),
    Reg("Output Active Power", 0x755E, 1, "u16",  1,     "W"),
    Reg("PV1 Input Power",     0x7540, 1, "u16",  1,     "W"),
    Reg("PV2 Input Power",     0x753D, 1, "u16",  1,     "W"),
    Reg("Battery Voltage",     0x7530, 1, "u16",  0.1,   "V"),
    Reg("Battery SOC",         0x7532, 1, "u16",  1,     "%"),
    Reg("Output Frequency",    0x754A, 1, "u16",  0.01,  "Hz"),
    Reg("Device Temperature",  0x7579, 1, "s16",  0.1,   "°C"),
]

SUMMARY_DATA: List[Reg] = [
    Reg("Line Charge Total",        0xCB61, 2, "u32", 0.0001, "kWh"),
    Reg("PV Generation Total",      0xCB56, 2, "u32", 0.0001, "kWh"),
    Reg("Load Consumption Total",   0xCB58, 2, "u32", 0.0001, "kWh"),
    Reg("Battery Charge Total",     0xCB52, 2, "u32", 0.0001, "kWh"),
    Reg("Battery Discharge Total",  0xCB54, 2, "u32", 0.0001, "kWh"),
    Reg("From Grid To Load",        0xCB63, 2, "u32", 0.0001, "kWh"),
    Reg("Operation Hours",          0xCBB0, 1, "u16", 1,      "h"),
]

ALL_REGS: Dict[int, Reg] = {r.addr: r for r in DEVICE_DATA + RUNTIME_DATA + SUMMARY_DATA}
REG_BY_NAME: Dict[str, Reg] = {r.name: r for r in DEVICE_DATA + RUNTIME_DATA + SUMMARY_DATA}

# ──────────────────────────────────────────────────────────────────────────────
# IDs

ID_EXIT                     = wx.NewId()
ID_SETTINGS                 = wx.NewId()
ID_TERM                     = wx.NewId()
ID_HELP                     = wx.NewId()

ID_PULL_ALL                 = wx.NewId()
ID_PULL_START               = wx.NewId()
ID_PULL_STOP                = wx.NewId()
ID_PULL_CLEAR               = wx.NewId()
ID_PULL_EXPORT              = wx.NewId()

# Keep “Send” menu items
ID_READ_SERIAL_NUMBER       = wx.NewId()
ID_READ_INVERTER_SN         = wx.NewId()
ID_READ_PRODUCTION_DATE     = wx.NewId()
ID_READ_FW                  = wx.NewId()
ID_READ_HW                  = wx.NewId()
ID_READ_MODEL_NUMBER        = wx.NewId()
ID_READ_MANUFACTURER        = wx.NewId()

ID_READ_AC_INPUT_VOLTAGE    = wx.NewId()
ID_READ_AC_INPUT_CURRENT    = wx.NewId()
ID_READ_AC_INPUT_POWER      = wx.NewId()
ID_READ_OUTPUT_ACTIVE_POWER = wx.NewId()
ID_READ_PV1_INPUT_POWER     = wx.NewId()
ID_READ_PV2_INPUT_POWER     = wx.NewId()
ID_READ_BATTERY_VOLTAGE     = wx.NewId()
ID_READ_BATTERY_SOC         = wx.NewId()
ID_READ_OUTPUT_FREQUENCY    = wx.NewId()
ID_READ_DEVICE_TEMPERATURE  = wx.NewId()

ID_READ_LINE_CHARGE_TOTAL       = wx.NewId()
ID_READ_PV_GENERATION_TOTAL     = wx.NewId()
ID_READ_LOAD_CONSUMPTION_TOTAL  = wx.NewId()
ID_READ_BATTERY_CHARGE_TOTAL    = wx.NewId()
ID_READ_BATTERY_DISCHARGE_TOTAL = wx.NewId()
ID_READ_FROM_GRID_TO_LOAD       = wx.NewId()
ID_READ_OPERATION_HOURS         = wx.NewId()

# Terminal settings
NEWLINE_CR, NEWLINE_LF, NEWLINE_CRLF = 0, 1, 2

class TerminalSetup:
    def __init__(self):
        self.echo = False
        self.unprintable = False
        self.newline = NEWLINE_CRLF

class TerminalSettingsDialog(wx.Dialog):
    def __init__(self, *args, **kwds):
        self.settings = kwds['settings']
        del kwds['settings']
        kwds["style"] = wx.DEFAULT_DIALOG_STYLE
        super().__init__(*args, **kwds)
        self.checkbox_echo = wx.CheckBox(self, -1, "Local Echo")
        self.checkbox_unprintable = wx.CheckBox(self, -1, "Show unprintable characters")
        self.radio_box_newline = wx.RadioBox(
            self, -1, "Newline Handling",
            choices=["CR only", "LF only", "CR+LF"], style=wx.RA_SPECIFY_ROWS
        )
        self.button_ok = wx.Button(self, -1, "OK")
        self.button_cancel = wx.Button(self, -1, "Cancel")
        self.SetTitle("Terminal Settings")
        self.button_ok.SetDefault()

        sizer_2 = wx.BoxSizer(wx.VERTICAL)
        sizer_3 = wx.BoxSizer(wx.HORIZONTAL)
        sizer_4 = wx.StaticBoxSizer(wx.StaticBox(self, -1, "Input/Output"), wx.VERTICAL)
        sizer_4.Add(self.checkbox_echo, 0, wx.ALL, 4)
        sizer_4.Add(self.checkbox_unprintable, 0, wx.ALL, 4)
        sizer_4.Add(self.radio_box_newline, 0, 0, 0)
        sizer_2.Add(sizer_4, 0, wx.EXPAND, 0)
        sizer_3.Add(self.button_ok, 0, 0, 0)
        sizer_3.Add(self.button_cancel, 0, 0, 0)
        sizer_2.Add(sizer_3, 0, wx.ALL | wx.ALIGN_RIGHT, 4)
        self.SetSizerAndFit(sizer_2)

        self.checkbox_echo.SetValue(self.settings.echo)
        self.checkbox_unprintable.SetValue(self.settings.unprintable)
        self.radio_box_newline.SetSelection(self.settings.newline)

        self.Bind(wx.EVT_BUTTON, self.OnOK, id=self.button_ok.GetId())
        self.Bind(wx.EVT_BUTTON, self.OnCancel, id=self.button_cancel.GetId())

    def OnOK(self, _):
        self.settings.echo = self.checkbox_echo.GetValue()
        self.settings.unprintable = self.checkbox_unprintable.GetValue()
        self.settings.newline = self.radio_box_newline.GetSelection()
        self.EndModal(wx.ID_OK)

    def OnCancel(self, _): self.EndModal(wx.ID_CANCEL)

# Menubar
class seWSNMenubar(wx.Frame):
    def __init__(self, parent):
        parent.seWSNView_menubar = wx.MenuBar()
        parent.SetMenuBar(parent.seWSNView_menubar)

        file_menu = wx.Menu()
        file_menu.Append(ID_EXIT, "&Exit", "")
        parent.Bind(wx.EVT_MENU, parent.OnExit, id=ID_EXIT)
        parent.seWSNView_menubar.Append(file_menu, "&File")

        config_menu = wx.Menu()
        config_menu.Append(ID_SETTINGS, "&Port Settings...", "")
        config_menu.Append(ID_TERM, "&Terminal Settings...", "")
        parent.Bind(wx.EVT_MENU, parent.OnPortSettings, id=ID_SETTINGS)
        parent.Bind(wx.EVT_MENU, parent.OnTermSettings, id=ID_TERM)
        parent.seWSNView_menubar.Append(config_menu, "&Config")

        send_menu = wx.Menu()
        send_menu.Append(ID_READ_SERIAL_NUMBER,       "Get Serial Numbers")
        send_menu.Append(ID_READ_INVERTER_SN,         "Get INVERTER SN")
        send_menu.Append(ID_READ_PRODUCTION_DATE,     "Get Production Date")
        send_menu.Append(ID_READ_FW,                  "Get Firmware Version")
        send_menu.Append(ID_READ_HW,                  "Get Hardware Version")
        send_menu.Append(ID_READ_MODEL_NUMBER,        "Get Model Number")
        send_menu.Append(ID_READ_MANUFACTURER,        "Get Manufacturer")
        send_menu.AppendSeparator()
        send_menu.Append(ID_READ_AC_INPUT_VOLTAGE,    "Get AC Input Voltage")
        send_menu.Append(ID_READ_AC_INPUT_CURRENT,    "Get AC Input Current")
        send_menu.Append(ID_READ_AC_INPUT_POWER,      "Get AC Input Power")
        send_menu.Append(ID_READ_OUTPUT_ACTIVE_POWER, "Get Output Active Power")
        send_menu.Append(ID_READ_PV1_INPUT_POWER,     "Get PV1 Input Power")
        send_menu.Append(ID_READ_PV2_INPUT_POWER,     "Get PV2 Input Power")
        send_menu.Append(ID_READ_BATTERY_VOLTAGE,     "Get Battery Voltage")
        send_menu.Append(ID_READ_BATTERY_SOC,         "Get Battery SOC")
        send_menu.Append(ID_READ_OUTPUT_FREQUENCY,    "Get Output Frequency")
        send_menu.Append(ID_READ_DEVICE_TEMPERATURE,  "Get Device Temperature")
        send_menu.AppendSeparator()
        send_menu.Append(ID_READ_LINE_CHARGE_TOTAL,       "Get Line Charge Total")
        send_menu.Append(ID_READ_PV_GENERATION_TOTAL,     "Get PV Generation Total")
        send_menu.Append(ID_READ_LOAD_CONSUMPTION_TOTAL,  "Get Load Consumption Total")
        send_menu.Append(ID_READ_BATTERY_CHARGE_TOTAL,    "Get Battery Charge Total")
        send_menu.Append(ID_READ_BATTERY_DISCHARGE_TOTAL, "Get Battery Discharge Total")
        send_menu.Append(ID_READ_FROM_GRID_TO_LOAD,       "Get From Grid To Load")
        send_menu.Append(ID_READ_OPERATION_HOURS,         "Get Operation Hours")
        parent.seWSNView_menubar.Append(send_menu, "&Send")

        help_menu = wx.Menu()
        help_menu.Append(ID_HELP, "&Help", "")
        parent.Bind(wx.EVT_MENU, parent.OnHelp, id=ID_HELP)
        parent.seWSNView_menubar.Append(help_menu, "&Help")

        pull_menu = wx.Menu()
        pull_menu.Append(ID_PULL_ALL,    "Pull all data once", "")
        pull_menu.AppendSeparator()
        pull_menu.Append(ID_PULL_START,  "Start", "")
        pull_menu.Append(ID_PULL_STOP,   "Stop", "")
        pull_menu.Append(ID_PULL_CLEAR,  "Clear", "")
        pull_menu.AppendSeparator()
        pull_menu.Append(ID_PULL_EXPORT, "Export data", "")
        parent.Bind(wx.EVT_MENU, parent.OnPullAll,     id=ID_PULL_ALL)
        parent.Bind(wx.EVT_MENU, parent.OnStartAuto,   id=ID_PULL_START)
        parent.Bind(wx.EVT_MENU, parent.OnStopAuto,    id=ID_PULL_STOP)
        parent.Bind(wx.EVT_MENU, parent.OnClearAll,    id=ID_PULL_CLEAR)
        parent.Bind(wx.EVT_MENU, parent.OnExportData,  id=ID_PULL_EXPORT)
        parent.seWSNView_menubar.Append(pull_menu, "Pull data")

# Pages
class PageTerminalView(wx.Panel):
    def __init__(self, parent):
        super().__init__(parent=parent, id=wx.ID_ANY)
        self.text_ctrl_output = wx.TextCtrl(self, wx.ID_ANY, "", style=wx.TE_MULTILINE | wx.TE_READONLY)
        s = wx.BoxSizer(wx.VERTICAL)
        s.Add(self.text_ctrl_output, 1, wx.EXPAND, 0)
        self.SetSizer(s)

class PageNetworkMonitor(wx.Panel):
    """Three clean columns at top; alarms box spans full width below."""
    def __init__(self, parent):
        super().__init__(parent=parent, id=wx.ID_ANY)
        self.SetDoubleBuffered(True)

        base_font: wx.Font = self.GetFont()
        label_font = wx.Font(pointSize=max(12, base_font.GetPointSize()),
                             family=base_font.GetFamily(),
                             style=wx.FONTSTYLE_NORMAL,
                             weight=wx.FONTWEIGHT_NORMAL)
        value_font = wx.Font(pointSize=max(12, base_font.GetPointSize()),
                             family=base_font.GetFamily(),
                             style=wx.FONTSTYLE_NORMAL,
                             weight=wx.FONTWEIGHT_NORMAL)

        tc_style = wx.TE_READONLY | wx.TE_RIGHT
        self.field_by_name: Dict[str, wx.TextCtrl] = {}
        self.field_by_addr: Dict[int, wx.TextCtrl] = {}

        # Top row: 3 columns
        top = wx.BoxSizer(wx.HORIZONTAL)

        def make_column(title: str, regs: List[Reg]):
            box = wx.StaticBox(self, wx.ID_ANY, title)
            col = wx.StaticBoxSizer(box, wx.VERTICAL)

            grid = wx.FlexGridSizer(rows=len(regs), cols=2, hgap=10, vgap=8)
            grid.AddGrowableCol(1, 1)
            for i in range(len(regs)):
                grid.AddGrowableRow(i, 1)

            def add_row(r: Reg):
                lbl = wx.StaticText(self, wx.ID_ANY, r.name)
                lbl.SetFont(label_font)
                txt = wx.TextCtrl(self, wx.ID_ANY, style=tc_style)
                txt.SetFont(value_font)
                txt.SetMinSize((240, 28))
                grid.Add(lbl, 0, wx.ALIGN_RIGHT | wx.ALIGN_CENTER_VERTICAL | wx.LEFT, 2)
                grid.Add(txt, 1, wx.EXPAND | wx.RIGHT, 2)
                self.field_by_name[r.name] = txt
                self.field_by_addr[r.addr] = txt

            for r in regs:
                add_row(r)

            col.Add(grid, 1, wx.EXPAND | wx.ALL, 8)
            return col

        top.Add(make_column("Device Data",  DEVICE_DATA), 1, wx.EXPAND | wx.ALL, 6)
        top.Add(make_column("Run-time Data", RUNTIME_DATA), 1, wx.EXPAND | wx.ALL, 6)
        top.Add(make_column("Summary Data", SUMMARY_DATA), 1, wx.EXPAND | wx.ALL, 6)

        # Bottom: alarms box (spans full width)
        fault_box = wx.StaticBox(self, wx.ID_ANY, "Active Alarms / Faults")
        fault_col = wx.StaticBoxSizer(fault_box, wx.VERTICAL)
        self.faults_text = wx.TextCtrl(
            self, wx.ID_ANY, "",
            style=wx.TE_MULTILINE | wx.TE_READONLY | wx.TE_RICH2
        )
        self.faults_text.SetMinSize((-1, 120))
        fault_col.Add(self.faults_text, 1, wx.EXPAND | wx.ALL, 8)

        root = wx.BoxSizer(wx.VERTICAL)
        root.Add(top, 1, wx.EXPAND)
        root.Add(fault_col, 0, wx.EXPAND | wx.LEFT | wx.RIGHT | wx.BOTTOM, 6)

        self.SetSizer(root)

# ──────────────────────────────────────────────────────────────────────────────
# NEW: Clean Machine Status page with controls
class PageMachinestatus(wx.Panel):
    """
    Clean control panel with rows:
      - ECO Mode ON/OFF
      - Generator Mode ON/OFF (LineRangeSet: OFF->UPS=0, ON->Generator=2)
      - Buzzer Mute ON/OFF
      - Battery SOC Low Shutdown (A09B) [read/enter/write]
      - Battery Full SOC Judgment (A09D) [read/enter/write]
    """
    def __init__(self, parent):
        super().__init__(parent=parent, id=wx.ID_ANY)
        self.SetDoubleBuffered(True)

        self.status_boxes: Dict[str, wx.TextCtrl] = {}
        self.spin_boxes: Dict[str, wx.SpinCtrl] = {}

        root = wx.BoxSizer(wx.VERTICAL)

        # ── Group 1: Operating modes
        modes_box = wx.StaticBox(self, wx.ID_ANY, "Operating Modes")
        modes = wx.StaticBoxSizer(modes_box, wx.VERTICAL)

        def add_toggle_row(label: str, key: str,
                           read_fn, set_on_fn, set_off_fn,
                           explain: Optional[str] = None):
            row = wx.BoxSizer(wx.HORIZONTAL)
            row.Add(wx.StaticText(self, label=label), 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 4)

            status = wx.TextCtrl(self, style=wx.TE_READONLY)
            status.SetMinSize((240, 28))
            self.status_boxes[key] = status
            row.Add(status, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 4)

            btn_read = wx.Button(self, label="Read")
            btn_on   = wx.Button(self, label="Set ON")
            btn_off  = wx.Button(self, label="Set OFF")
            row.Add(btn_read, 0, wx.ALL, 2)
            row.Add(btn_on,   0, wx.ALL, 2)
            row.Add(btn_off,  0, wx.ALL, 2)

            if explain:
                row.Add(wx.StaticText(self, label=f"  {explain}"), 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 2)

            btn_read.Bind(wx.EVT_BUTTON, lambda _e: read_fn(key))
            btn_on.Bind(wx.EVT_BUTTON,   lambda _e: set_on_fn(key))
            btn_off.Bind(wx.EVT_BUTTON,  lambda _e: set_off_fn(key))

            modes.Add(row, 0, wx.ALL, 2)

        add_toggle_row(
            "ECO Mode:", "eco",
            self._read_eco, self._set_eco_on, self._set_eco_off,
            # "A02D EcoEn 0:OFF / 1:ON"
        )
        add_toggle_row(
            "Generator Mode:", "gen",
            self._read_gen, self._set_gen_on, self._set_gen_off,
            # "A02B LineRangeSet 0:UPS 1:APL 2:GEN"
        )
        add_toggle_row(
            "Buzzer Mute:", "mute",
            self._read_mute, self._set_mute_on, self._set_mute_off,
            # "A033 MuteEn 0:OFF / 1:ON"
        )

        root.Add(modes, 0, wx.EXPAND | wx.ALL, 6)

        # ── Group 2: SOC settings
        soc_box = wx.StaticBox(self, wx.ID_ANY, "Battery SOC Settings")
        soc = wx.StaticBoxSizer(soc_box, wx.VERTICAL)

        def add_soc_row(label: str, key: str, read_fn, write_fn):
            row = wx.BoxSizer(wx.HORIZONTAL)
            row.Add(wx.StaticText(self, label=label), 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 4)

            status = wx.TextCtrl(self, style=wx.TE_READONLY)
            status.SetMinSize((120, 28))
            self.status_boxes[key] = status
            row.Add(status, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 4)

            sp = wx.SpinCtrl(self, min=0, max=100, initial=0)
            sp.SetMinSize((90, 28))
            self.spin_boxes[key] = sp
            row.Add(wx.StaticText(self, label=" Set: "), 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 2)
            row.Add(sp, 0, wx.ALL | wx.ALIGN_CENTER_VERTICAL, 2)

            btn_read  = wx.Button(self, label="Read")
            btn_write = wx.Button(self, label="Write")
            row.Add(btn_read, 0, wx.ALL, 2)
            row.Add(btn_write, 0, wx.ALL, 2)

            btn_read.Bind(wx.EVT_BUTTON, lambda _e: read_fn(key))
            btn_write.Bind(wx.EVT_BUTTON, lambda _e: write_fn(key))

            soc.Add(row, 0, wx.ALL, 2)

        add_soc_row("Low shutdown SOC:", "soc_stop", self._read_soc_stop, self._write_soc_stop)
        add_soc_row("Full SOC judgment:", "soc_full", self._read_soc_full, self._write_soc_full)

        root.Add(soc, 0, wx.EXPAND | wx.ALL, 6)

        self.SetSizer(root)

        # Auto-refresh once when tab is created
        wx.CallAfter(self._read_eco, "eco")
        wx.CallAfter(self._read_gen, "gen")
        wx.CallAfter(self._read_mute, "mute")
        wx.CallAfter(self._read_soc_stop, "soc_stop")
        wx.CallAfter(self._read_soc_full, "soc_full")

    # helpers to reach the frame
    def _frm(self): return self.GetTopLevelParent()

    # --------------- ECO ----------------
    def _read_eco(self, key):
        v = self._frm().mb_read_u16(0xA02D)
        self._set_status_text(key, "ON" if v == 1 else "OFF" if v is not None else "—")

    def _set_eco_on(self, key):
        if self._frm().mb_write_single(0xA02D, 1): self._read_eco(key)

    def _set_eco_off(self, key):
        if self._frm().mb_write_single(0xA02D, 0): self._read_eco(key)

    # ------------- Generator -------------
    def _read_gen(self, key):
        v = self._frm().mb_read_u16(0xA02B)
        txt = {0: "UPS (OFF)", 1: "APL", 2: "GEN (ON)"}.get(v, str(v)) if v is not None else "—"
        self._set_status_text(key, txt)

    def _set_gen_on(self, key):
        if self._frm().mb_write_single(0xA02B, 2): self._read_gen(key)  # 2=Generator

    def _set_gen_off(self, key):
        if self._frm().mb_write_single(0xA02B, 0): self._read_gen(key)  # 0=UPS (treat as OFF)

    # --------------- Mute ----------------
    def _read_mute(self, key):
        v = self._frm().mb_read_u16(0xA033)
        self._set_status_text(key, "ON" if v == 1 else "OFF" if v is not None else "—")

    def _set_mute_on(self, key):
        if self._frm().mb_write_single(0xA033, 1): self._read_mute(key)

    def _set_mute_off(self, key):
        if self._frm().mb_write_single(0xA033, 0): self._read_mute(key)

    # --------------- SOC -----------------
    def _read_soc_stop(self, key):
        v = self._frm().mb_read_u16(0xA09B)
        if v is not None:
            self._set_status_text(key, f"{v} %")
            self.spin_boxes[key].SetValue(int(v))

    def _write_soc_stop(self, key):
        val = int(self.spin_boxes[key].GetValue())
        if self._frm().mb_write_single(0xA09B, val): self._read_soc_stop(key)

    def _read_soc_full(self, key):
        v = self._frm().mb_read_u16(0xA09D)
        if v is not None:
            self._set_status_text(key, f"{v} %")
            self.spin_boxes[key].SetValue(int(v))

    def _write_soc_full(self, key):
        val = int(self.spin_boxes[key].GetValue())
        if self._frm().mb_write_single(0xA09D, val): self._read_soc_full(key)

    # Common UI helper
    def _set_status_text(self, key: str, text: str):
        box = self.status_boxes.get(key)
        if box:
            box.SetValue(text)

# ──────────────────────────────────────────────────────────────────────────────
# Main window
class seWSNViewLayout(wx.Frame):
    POLL_SECONDS_DEFAULT = 10

    # Popup behavior controls
    _NOT_CONNECTED_GRACE_S = 6.0   # don't show popup during the first N seconds
    _NOT_CONNECTED_COOLDOWN_S = 3.0  # show at most once every N seconds

    # Alarm descriptions (partial)
    FAULT_DESC: Dict[int, str] = {
        1: "Battery under voltage warning",
        2: "Battery under voltage protection ",
        3: "Average battery discharge current over current protection",
        4: "Instantaneous battery discharge over current protection",
        5: "Battery not connected ",
        6: "Battery over voltage ",
        7: "BMS low battery alarm",
        8: "BMS low battery protection",
        9: "Bypass overload protection",
        10: "Battery output overload protection",
        11: "Battery inverter output short circuit",
        12: "The AC output of the battery inverter over circuit",
        13: "The DC component of the battery inverter voltage is abnormal",
        14: "Bus over voltage software sampling protection",
        15: "Bus over voltage hardware sampling protection",
        16: "Bus under voltage protection",
        17: "Bus short circuit protection",
        18: "The PV input voltage is over voltage",
        20: "PV over current protection",
        22: "The PV heat sink is overheated",
        23: "The AC heat sink is overheated.",
        24: "The temperature of the main transformer is overheated",
        25: "Ac input relay short circuit",
        27: "Fan Failure",
        30: "Type detection error",
        33: "Parallel control can communication is faulty",
        34: "Parallel control can communication is faulty",
        35: "Parallel mode is faulty ",
        36: "Parallel current sharing fault",
        37: "Parallel ID setting error",
        38: "Inconsistent Battery in parallel mode",
        39: "Inconsistent AC input source in parallel mode",
        40: "The parallel mode synchronization fails",
        41: "Inconsistent system firmware version in parallel mode",
        42: "The parallel communication cable is faulty",
        43: "Serial number error",
        49: "BMS communication error",
        50: "BMS other alarm",
        51: "BMS battery over temperature",
        52: "BMS battery over current",
        53: "BMS battery over voltage",
        54: "BMS battery low voltage",
        55: "BMS battery low temperature",
        56: "PD communication error",
        58: "BMS pack number mismatch",
    }

    def __init__(self, *args, **kwds):
        super().__init__(*args, **kwds)
        self.SetTitle("REON Modbus GUI")
        self.SetSize((1300, 900))

        self.serial = serial.Serial()
        self.serial.timeout = 1.0

        self.settings = TerminalSetup()
        self.mb: Optional[ModbusSerialClient] = None
        self.mb_lock = threading.Lock()
        self.modbus_slave_id = 1

        self.poll_timer = wx.Timer(self)
        self.poll_period_ms = self.POLL_SECONDS_DEFAULT * 500
        self.Bind(wx.EVT_TIMER, self._on_poll_timer, self.poll_timer)

        # popup timing state
        self._app_started_at = time.time()
        self._last_not_connected_popup = 0.0

        seWSNMenubar(self)

        # ── Top-level panel holding a header (logo) + the notebook
        p = wx.Panel(self)

        # Header bar with centered logo
        self.header = wx.Panel(p)
        header_sizer = wx.BoxSizer(wx.HORIZONTAL)
        header_sizer.AddStretchSpacer(1)
        bmp = self._load_logo_bitmap(height_px=28)
        self.logo_ctrl = wx.StaticBitmap(self.header, bitmap=bmp)
        header_sizer.Add(self.logo_ctrl, 0, wx.ALL | wx.ALIGN_CENTER, 6)
        header_sizer.AddStretchSpacer(1)
        self.header.SetSizer(header_sizer)
        self.header.SetMinSize((-1, bmp.GetHeight() + 10))

        # Main notebook
        self.nb = wx.Notebook(p)
        self.pageNetMon = PageNetworkMonitor(self.nb)
        self.pageMachineStatus = PageMachinestatus(self.nb)  # clean control page
        self.pageTerminal = PageTerminalView(self.nb)
        self.nb.AddPage(self.pageNetMon, "Machine Monitor")
        self.nb.AddPage(self.pageMachineStatus, "Machine Configuration")
        self.nb.AddPage(self.pageTerminal, "Terminal View")
        self._set_notebook_tab_font(point_size_increase=6)

        # Layout: header on top, notebook fills the rest
        root_v = wx.BoxSizer(wx.VERTICAL)
        root_v.Add(self.header, 0, wx.EXPAND)
        root_v.Add(self.nb, 1, wx.EXPAND)
        p.SetSizer(root_v)

        # Bind Send menu items to generic readers by name
        b = self.Bind
        b(wx.EVT_MENU, lambda e: self.read_and_show("Serial #"),               id=ID_READ_SERIAL_NUMBER)
        b(wx.EVT_MENU, lambda e: self.read_and_show("Inverter SN"),            id=ID_READ_INVERTER_SN)
        b(wx.EVT_MENU, lambda e: self.read_and_show("Production Date"),        id=ID_READ_PRODUCTION_DATE)
        b(wx.EVT_MENU, lambda e: self.read_and_show("Firmware Version"),       id=ID_READ_FW)
        b(wx.EVT_MENU, lambda e: self.read_and_show("HW Version"),             id=ID_READ_HW)
        b(wx.EVT_MENU, lambda e: self.read_and_show("Model Number"),           id=ID_READ_MODEL_NUMBER)
        b(wx.EVT_MENU, lambda e: self.read_and_show("Manufacturer"),           id=ID_READ_MANUFACTURER)
        b(wx.EVT_MENU, lambda e: self.read_and_show("AC Input Voltage"),       id=ID_READ_AC_INPUT_VOLTAGE)
        b(wx.EVT_MENU, lambda e: self.read_and_show("AC Input Current"),       id=ID_READ_AC_INPUT_CURRENT)
        b(wx.EVT_MENU, lambda e: self.read_and_show("AC Input Power"),         id=ID_READ_AC_INPUT_POWER)
        b(wx.EVT_MENU, lambda e: self.read_and_show("Output Active Power"),    id=ID_READ_OUTPUT_ACTIVE_POWER)
        b(wx.EVT_MENU, lambda e: self.read_and_show("PV1 Input Power"),        id=ID_READ_PV1_INPUT_POWER)
        b(wx.EVT_MENU, lambda e: self.read_and_show("PV2 Input Power"),        id=ID_READ_PV2_INPUT_POWER)
        b(wx.EVT_MENU, lambda e: self.read_and_show("Battery Voltage"),        id=ID_READ_BATTERY_VOLTAGE)
        b(wx.EVT_MENU, lambda e: self.read_and_show("Battery SOC"),            id=ID_READ_BATTERY_SOC)
        b(wx.EVT_MENU, lambda e: self.read_and_show("Output Frequency"),       id=ID_READ_OUTPUT_FREQUENCY)
        b(wx.EVT_MENU, lambda e: self.read_and_show("Device Temperature"),     id=ID_READ_DEVICE_TEMPERATURE)
        b(wx.EVT_MENU, lambda e: self.read_and_show("Line Charge Total"),      id=ID_READ_LINE_CHARGE_TOTAL)
        b(wx.EVT_MENU, lambda e: self.read_and_show("PV Generation Total"),    id=ID_READ_PV_GENERATION_TOTAL)
        b(wx.EVT_MENU, lambda e: self.read_and_show("Load Consumption Total"), id=ID_READ_LOAD_CONSUMPTION_TOTAL)
        b(wx.EVT_MENU, lambda e: self.read_and_show("Battery Charge Total"),   id=ID_READ_BATTERY_CHARGE_TOTAL)
        b(wx.EVT_MENU, lambda e: self.read_and_show("Battery Discharge Total"),id=ID_READ_BATTERY_DISCHARGE_TOTAL)
        b(wx.EVT_MENU, lambda e: self.read_and_show("From Grid To Load"),      id=ID_READ_FROM_GRID_TO_LOAD)
        b(wx.EVT_MENU, lambda e: self.read_and_show("Operation Hours"),        id=ID_READ_OPERATION_HOURS)

        self.Bind(wx.EVT_CLOSE, self.OnClose)
        wx.CallAfter(self.autodetect_usb_and_connect)

    def _set_notebook_tab_font(self, point_size_increase=3):
        f = self.nb.GetFont()
        f.SetPointSize(f.GetPointSize() + int(point_size_increase))
        self.nb.SetFont(f)
        self.nb.Layout()
        self.nb.Refresh()
        self.GetChildren()[0].Layout()

    # Load and scale a logo; fall back to a stock bitmap if not found
    def _load_logo_bitmap(self, height_px: int = 28) -> wx.Bitmap:
        here = os.path.dirname(os.path.abspath(__file__))
        for name in ("company_logo.png", "logo.png", "logo.jpg", "logo.bmp"):
            path = os.path.join(here, name)
            if os.path.exists(path):
                img = wx.Image(path)
                if img.IsOk():
                    w = max(1, int(img.GetWidth() * (height_px / float(img.GetHeight()))))
                    img = img.Scale(w, height_px, wx.IMAGE_QUALITY_HIGH)
                    return wx.Bitmap(img)
        return wx.ArtProvider.GetBitmap(wx.ART_INFORMATION, wx.ART_OTHER, (height_px, height_px))

    # UI helpers
    def UpdatePageTerminal(self, s):
        try:
            self.pageTerminal.text_ctrl_output.AppendText(str(s))
        except Exception:
            pass

    def OnExit(self, _): self.Close()

    def OnClose(self, _):
        try:
            if self.mb:
                try:
                    if getattr(self.mb, "connected", False):
                        self.mb.close()
                except Exception:
                    self.mb.close()
        except Exception:
            pass
        self.poll_timer.Stop()
        self.Destroy()

    def OnHelp(self, _):
        message = (
            "Version Information:\n\n"
            f"pymodbus: {getattr(pymodbus, '__version__', '?')}\n"
            f"Framer:   {_FRAMER_KW}\n"
            "Comments: Engineering build (Modbus RTU)\n"
        )
        wx.MessageBox(message, "Help About", wx.OK | wx.ICON_INFORMATION)

    def OnTermSettings(self, _):
        dlg = TerminalSettingsDialog(None, -1, "", settings=self.settings)
        dlg.ShowModal()
        dlg.Destroy()

    # Modbus setup
    def _parity_char(self, pyserial_parity):
        try:
            from serial import PARITY_NONE, PARITY_EVEN, PARITY_ODD
            if pyserial_parity == PARITY_NONE: return 'N'
            if pyserial_parity == PARITY_EVEN: return 'E'
            if pyserial_parity == PARITY_ODD:  return 'O'
        except Exception:
            pass
        return str(pyserial_parity or 'N')

    def mb_connect_from_current_settings(self):
        try:
            if self.mb:
                try:
                    if getattr(self.mb, "connected", False):
                        self.mb.close()
                except Exception:
                    self.mb.close()
        except Exception:
            pass

        port_str = getattr(self.serial, "portstr", None) or getattr(self.serial, "port", None)

        if _HAS_FRAMER:
            self.mb = ModbusSerialClient(
                port=port_str,
                framer=FramerType.RTU,
                baudrate=self.serial.baudrate,
                bytesize=self.serial.bytesize,
                parity=self._parity_char(self.serial.parity),
                stopbits=self.serial.stopbits,
                timeout=self.serial.timeout or 1.0,
            )
        else:
            self.mb = ModbusSerialClient(
                method="rtu",
                port=port_str,
                baudrate=self.serial.baudrate,
                bytesize=self.serial.bytesize,
                parity=self._parity_char(self.serial.parity),
                stopbits=self.serial.stopbits,
                timeout=self.serial.timeout or 1.0,
            )

        ok = self.mb.connect()
        return bool(ok)

    def OnPortSettings(self, _=None):
        try:
            dlg = wxSerialConfigDialog.SerialConfigDialog(
                self, -1, "",
                show=(wxSerialConfigDialog.SHOW_BAUDRATE
                      | wxSerialConfigDialog.SHOW_FORMAT
                      | wxSerialConfigDialog.SHOW_FLOW),
                serial=self.serial
            )
            if dlg.ShowModal() == wx.ID_OK:
                if self.mb_connect_from_current_settings():
                    self._update_title_connected()
                    self.UpdatePageTerminal("Modbus RTU connected.\n")
                    self.UpdatePageTerminal(f"pymodbus {getattr(pymodbus, '__version__', '?')} using {_FRAMER_KW}\n")
                else:
                    wx.MessageBox("Failed to connect via Modbus RTU with the selected settings.",
                                  "Connection Error", wx.OK | wx.ICON_ERROR)
            dlg.Destroy()
        except Exception as e:
            wx.MessageBox(f"Error in port settings: {e}", "Error", wx.OK | wx.ICON_ERROR)

    def _update_title_connected(self):
        port_label = getattr(self.serial, "portstr", None) or getattr(self.serial, "port", "")
        self.SetTitle(
            f"REON Modbus GUI tool on {port_label} "
            f"[{self.serial.baudrate},{self.serial.bytesize}{self._parity_char(self.serial.parity)}{self.serial.stopbits}]"
        )

    # Auto-detect + connect
    def _choose_usb_port(self, ports):
        candidates = []
        for p in ports:
            dev = (p.device or "").lower()
            desc = (p.description or "").lower()
            looks_usb = (
                "usb" in desc or
                any(x in dev for x in ["ttyusb", "ttyacm", "usbserial", "usbmodem"]) or
                getattr(p, "vid", None) is not None
            )
            if not looks_usb or "bluetooth" in desc:
                continue
            score = 0
            if any(x in dev for x in ["ttyusb", "ttyacm", "usbserial", "usbmodem"]): score += 50
            if any(x in desc for x in ["ftdi", "cp210", "ch340", "ch341", "prolific", "silicon labs", "cdc", "usb serial"]): score += 30
            if getattr(p, "vid", None) is not None: score += 10
            candidates.append((score, p))
        if not candidates:
            return None
        candidates.sort(key=lambda t: t[0], reverse=True)
        return candidates[0][1]

    def autodetect_usb_and_connect(self):
        try:
            ports = list(list_ports.comports())
        except Exception as e:
            self.UpdatePageTerminal(f"Auto-detect: list_ports error: {e}\n")
            return
        if not ports:
            self.UpdatePageTerminal("Auto-detect: no serial ports found.\n")
            return
        cand = self._choose_usb_port(ports)
        if not cand:
            self.UpdatePageTerminal("Auto-detect: no USB serial device found.\n")
            return

        self.serial.port     = cand.device
        self.serial.baudrate = getattr(self.serial, "baudrate", 9600) or 9600
        self.serial.bytesize = getattr(self.serial, "bytesize", 8) or 8
        self.serial.parity   = getattr(self.serial, "parity", serial.PARITY_NONE) or serial.PARITY_NONE
        self.serial.stopbits = getattr(self.serial, "stopbits", serial.STOPBITS_ONE) or serial.STOPBITS_ONE

        self.UpdatePageTerminal(f"Auto-detect: using {cand.device} ({cand.description})\n")

        if self.mb_connect_from_current_settings():
            self._update_title_connected()
            self.UpdatePageTerminal("Modbus RTU connected (auto-detected).\n")
            self.UpdatePageTerminal(f"pymodbus {getattr(pymodbus, '__version__', '?')} using {_FRAMER_KW}\n")
        else:
            self.UpdatePageTerminal("Auto-detect: failed to connect. Use Config → Port Settings…\n")

    # ── Not-connected popup helpers ───────────────────────────────────────────
    def _maybe_warn_not_connected(self):
        now = time.time()
        # Respect startup grace period and cooldown
        if (now - self._app_started_at) < self._NOT_CONNECTED_GRACE_S:
            return
        if (now - self._last_not_connected_popup) < self._NOT_CONNECTED_COOLDOWN_S:
            return
        self._last_not_connected_popup = now
        wx.MessageBox("Modbus client is not connected.", "Error", wx.OK | wx.ICON_ERROR)

    # ── Modbus read/write wrappers ────────────────────────────────────────────
    def _call_read(self, method_name, address, count, unit):
        if not self.mb:
            return None
        fn = getattr(self.mb, method_name, None)
        if not fn:
            return None
        try:
            return fn(address=address, count=count, slave=unit)
        except TypeError:
            pass
        try:
            return fn(address=address, count=count, unit=unit)
        except TypeError:
            pass
        try:
            return fn(address=address, count=count)
        except TypeError:
            pass
        try:
            return fn(address=address)
        except Exception:
            return None

    def _read_holding(self, address, count=1, unit=None):
        unit = unit or self.modbus_slave_id
        return self._call_read("read_holding_registers", address, count, unit)

    def mb_read_holding(self, address, count=1, unit=None):
        if not self.mb:
            self._maybe_warn_not_connected()
            return None
        with self.mb_lock:
            rr = self._read_holding(address, count, unit)
        if rr is None:
            self.UpdatePageTerminal(f"Read failed (None) at 0x{address:04X}\n"); return None
        try:
            if rr.isError():
                self.UpdatePageTerminal(f"Modbus error on 0x{address:04X}: {rr}\n"); return None
        except Exception:
            pass
        if getattr(rr, "registers", None) is None:
            self.UpdatePageTerminal(f"No data returned at 0x{address:04X}: {rr}\n"); return None
        return rr.registers

    def mb_read_u16(self, address, unit=None) -> Optional[int]:
        regs = self.mb_read_holding(address, 1, unit)
        if regs is None:
            return None
        return int(regs[0] & 0xFFFF)

    # write single (FC=06)
    def mb_write_single(self, address, value, unit=None) -> bool:
        if not self.mb:
            self._maybe_warn_not_connected()
            return False
        unit = unit or self.modbus_slave_id
        try:
            fn = getattr(self.mb, "write_register", None)
            if not fn:
                self.UpdatePageTerminal("write_register not available on Modbus client.\n")
                return False
            try:
                rr = fn(address=address, value=int(value) & 0xFFFF, slave=unit)
            except TypeError:
                try:
                    rr = fn(address=address, value=int(value) & 0xFFFF, unit=unit)
                except TypeError:
                    rr = fn(address=address, value=int(value) & 0xFFFF)
        except Exception as e:
            self.UpdatePageTerminal(f"Write error at 0x{address:04X}: {e}\n")
            return False

        try:
            if rr is None or rr.isError():
                self.UpdatePageTerminal(f"Modbus write error at 0x{address:04X}: {rr}\n")
                return False
        except Exception:
            pass
        self.UpdatePageTerminal(f"Wrote 0x{int(value) & 0xFFFF:04X} to 0x{address:04X}\n")
        return True

    # Decoders / formatters
    def _u16(self, w):  return int(w & 0xFFFF)
    def _s16(self, w):
        v = int(w & 0xFFFF)
        return v - 0x10000 if v & 0x8000 else v

    def _u32_be(self, hi, lo): return ((hi & 0xFFFF) << 16) | (lo & 0xFFFF)
    def _s32_be(self, hi, lo):
        u = self._u32_be(hi, lo)
        return u - 0x1_0000_0000 if u & 0x8000_0000 else u

    def _decode(self, regs, codec):
        if not regs: return None
        if codec == "ascii":
            b = b"".join(int(r & 0xFFFF).to_bytes(2, "big") for r in regs)
            return b.split(b"\x00", 1)[0].decode("ascii", errors="ignore")
        if codec == "u16": return self._u16(regs[0])
        if codec == "s16": return self._s16(regs[0])
        if codec == "u32":
            if len(regs) < 2: return None
            return self._u32_be(regs[0], regs[1])
        if codec == "s32":
            if len(regs) < 2: return None
            return self._s32_be(regs[0], regs[1])
        return None

    def _fmt_scaled(self, val, scale: float, unit: str) -> str:
        if val is None: return ""
        if isinstance(val, (int, float)) and scale != 1:
            from decimal import Decimal
            dec = max(0, -Decimal(str(scale)).as_tuple().exponent)
            s = f"{val * scale:.{dec}f}"
        elif isinstance(val, (int, float)):
            s = f"{val:.0f}"
        else:
            s = str(val)
        return f"{s} {unit}".strip()

    # ---- Active alarm helpers ----
    def _read_active_alarm_ids(self) -> List[int]:
        # 8x16 bits starting at 0x75A5 => alarms 1..128
        regs = self.mb_read_holding(0x75A5, 8)
        if regs is None:
            return []
        ids: List[int] = []
        for w_idx, w in enumerate(regs):
            for bit in range(16):
                if w & (1 << bit):
                    ids.append(w_idx * 16 + bit + 1)
        return ids

    def _update_alarm_box(self):
        ids = self._read_active_alarm_ids()
        if not hasattr(self.pageNetMon, "faults_text"):
            return
        if not ids:
            self.pageNetMon.faults_text.SetValue("No active alarms.")
            return
        lines = []
        for a in ids:
            label = self.FAULT_DESC.get(a, f"Alarm {a}")
            detail = self.mb_read_holding(0x9A4C + (a - 1), 1)
            if detail is not None and len(detail) == 1 and detail[0] != 0:
                label += f" (detail={detail[0]})"
            lines.append(f"[{a:02d}] {label}")
        self.pageNetMon.faults_text.SetValue("\n".join(lines))

    # Generic read + show
    def read_and_show(self, reg_name: str):
        reg = REG_BY_NAME.get(reg_name)
        if not reg:
            self.UpdatePageTerminal(f"Unknown register '{reg_name}'\n")
            return
        regs = self.mb_read_holding(reg.addr, reg.words)
        if regs is None:
            return
        decoded = self._decode(regs, reg.codec)
        text = self._fmt_scaled(decoded, reg.scale, reg.unit) if reg.codec != "ascii" else str(decoded)
        ctrl = self.pageNetMon.field_by_name.get(reg.name)
        if ctrl:
            ctrl.SetValue(text)
        self.UpdatePageTerminal(f"{reg.name}: {text}\n")

    # Batch: Pull all data once
    def OnPullAll(self, _):
        if not self.mb:
            self._maybe_warn_not_connected()
            return
        self.UpdatePageTerminal("Pulling all data...\n")
        for reg in DEVICE_DATA + RUNTIME_DATA + SUMMARY_DATA:
            try:
                self.read_and_show(reg.name)
                time.sleep(0.02)
            except Exception as e:
                self.UpdatePageTerminal(f"Error during {reg.name}: {e}\n")
        self._update_alarm_box()
        self.UpdatePageTerminal("Done pulling all data.\n")

    # Start/Stop/Clear
    def OnStartAuto(self, _=None):
        if not self.poll_timer.IsRunning():
            self.poll_timer.Start(self.poll_period_ms)
            self.UpdatePageTerminal(f"Auto-poll started ({self.poll_period_ms/1000:.0f}s).\n")

    def OnStopAuto(self, _=None):
        if self.poll_timer.IsRunning():
            self.poll_timer.Stop()
            self.UpdatePageTerminal("Auto-poll stopped.\n")

    def _on_poll_timer(self, _evt):
        self.OnPullAll(None)

    def OnClearAll(self, _=None):
        for ctrl in self.pageNetMon.field_by_name.values():
            ctrl.SetValue("")
        if hasattr(self.pageNetMon, "faults_text"):
            self.pageNetMon.faults_text.SetValue("")
        self.UpdatePageTerminal("Cleared all fields.\n")

    # Excel export
    def OnExportData(self, _=None):
        try:
            from openpyxl import Workbook
        except ImportError:
            wx.MessageBox(
                "The 'openpyxl' package is required to export Excel files.\n"
                "Install with:  pip install openpyxl",
                "Missing dependency", wx.OK | wx.ICON_ERROR
            )
            return

        sections = [
            ("Device Data", DEVICE_DATA),
            ("Run-time Data", RUNTIME_DATA),
            ("Summary Data", SUMMARY_DATA),
        ]
        rows = [("Name", "Value")]
        for title, reg_list in sections:
            rows.append((title, ""))
            for reg in reg_list:
                ctrl = self.pageNetMon.field_by_name.get(reg.name)
                rows.append((reg.name, ctrl.GetValue() if ctrl else ""))
            rows.append(("", ""))

        dlg = wx.FileDialog(
            self, "Save data as",
            wildcard="Excel files (*.xlsx)|*.xlsx",
            style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT
        )
        if dlg.ShowModal() != wx.ID_OK:
            dlg.Destroy()
            return
        path = dlg.GetPath()
        dlg.Destroy()

        wb = Workbook()
        ws = wb.active
        ws.title = "Snapshot"
        for r, (name, value) in enumerate(rows, start=1):
            ws.cell(row=r, column=1, value=name)
            ws.cell(row=r, column=2, value=value)
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 40
        ws.freeze_panes = "A2"
        wb.save(path)
        self.UpdatePageTerminal(f"Exported {len(rows)-1} rows to {path}\n")

# App
class MyApp(wx.App):
    def OnInit(self):
        self.frame = seWSNViewLayout(None, -1, "")
        self.SetTopWindow(self.frame)
        self.frame.Show()
        return True

if __name__ == "__main__":
    app = MyApp(False)
    app.MainLoop()
